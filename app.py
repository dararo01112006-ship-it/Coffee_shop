import os
import random
import requests
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory, render_template_string, session, redirect, url_for
from geopy.distance import geodesic
from openpyxl import Workbook, load_workbook
from functools import wraps

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = "your-secret-key-change-this-in-production"  # ⚠️ សូមប្តូរនៅពេល Deploy

# ================== ការកំណត់ (CONFIG) ==================
BOT_TOKEN = "8501341500:AAFvNtQIAzELusb_5u6EPgSjGMpBcv0avpo"
CHAT_ID = 8091370821
 
SHOP_LAT = 11.530131512325177
SHOP_LON = 104.8850590846574
MAX_DISTANCE = 2000  # ម៉ែត្រ
EXCEL_FILE = "orders.xlsx"
# Admin Login Config
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "1234"  # ⚠️ សូមប្តូរ password នេះ!

# =================== HELPER FUNCTIONS ===================
def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(["Order ID", "Queue Number", "Time", "Items", "Total ($)", "Distance (m)", "Map Link"])
        wb.save(EXCEL_FILE)
        print("✅ Excel file initialized.")

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function

init_excel()

# ================== ROUTES =====================

@app.route("/")
def index():
    return send_from_directory('.', 'testweb.html')

@app.route("/order", methods=["POST"])
def order():
    try:
        data = request.json
        if not data:
            return jsonify({"success": False, "error": "មិនមានទិន្នន័យបញ្ជូនមក"}), 400

        # ទាញយកទិន្នន័យ
        queue_number = data.get("queueNumber", "N/A")
        items = data.get("items", [])
        total = data.get("total", 0)
        location = data.get("location", {})
        lat = location.get("lat")
        lon = location.get("lng")

        # 1. បង្កើតព័ត៌មានបឋម
        order_id = f"ORD{random.randint(1000, 9999)}"
        time_now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # 2. ពិនិត្យចម្ងាយ (Distance Logic) - STRICT CHECK
        distance_m = None
        distance_text = "មិនបានបើក GPS"
        map_link = "N/A"

        # ⚠️ STRICT: ត្រូវតែមាន GPS location
        if not lat or not lon:
            return jsonify({
                "success": False, 
                "error": "សូមបើក GPS និងអនុញ្ញាតឲ្យ Browser ចូលប្រើទីតាំងរបស់អ្នក!"
            }), 400

        # គណនាចម្ងាយ
        user_coords = (lat, lon)
        shop_coords = (SHOP_LAT, SHOP_LON)
        distance_m = geodesic(shop_coords, user_coords).meters
        
        # ⚠️ STRICT: បដិសេធភ្លាមប្រសិនលើសចម្ងាយកំណត់
        if distance_m > MAX_DISTANCE:
            return jsonify({
                "success": False, 
                "error": f"❌ សូមអភ័យទោស!\n\nអ្នកនៅឆ្ងាយពីហាងពេក!\n\nចម្ងាយរបស់អ្នក: {round(distance_m)} ម៉ែត្រ\nចម្ងាយអតិបរមា: {MAX_DISTANCE} ម៉ែត្រ\n\nយើងមិនអាចដឹកជញ្ជូនទៅកាន់ទីតាំងរបស់អ្នកបានទេ។"
            }), 403
        
        distance_text = f"{round(distance_m, 2)} ម៉ែត្រ"
        map_link = f"https://www.google.com/maps?q={lat},{lon}"

        # 3. រៀបចំបញ្ជីមុខម្ហូប (Item Details)
        if not items:
            return jsonify({"success": False, "error": "សូមជ្រើសរើសមុខម្ហូបមុននឹងកុម្ម៉ង់"}), 400

        items_detail_msg = ""  # សម្រាប់ Telegram
        items_for_excel = ""   # សម្រាប់ Excel
        
        for item in items:
            name = item.get('name_km') or item.get('name_en') or 'Unknown'
            qty = item.get('qty', 1)
            price = item.get('price', 0)
            subtotal = price * qty
            
            # ជម្រើសបន្ថែម (Options)
            opts = []
            if item.get('options'):
                opt = item['options']
                if opt.get('sugar'): opts.append(f"ស្ករ:{opt['sugar']}")
                if opt.get('ice'): opts.append(f"ទឹកកក:{opt['ice']}")
                if opt.get('note'): opts.append(f"ចំណាំ:{opt['note']}")
            
            opt_str = f" ({', '.join(opts)})" if opts else ""
            
            items_detail_msg += f"• {name} x{qty}{opt_str} = ${subtotal:.2f}\n"
            items_for_excel += f"{name}(x{qty}){opt_str}, "

        # 4. រៀបចំសារផ្ញើទៅ Telegram
        telegram_msg = (
            f"🔔 **មានការកុម្ម៉ង់ថ្មី!**\n\n"
            f"🎫 លេខរង់ចាំ: `{queue_number}`\n"
            f"🆔 លេខកុម្ម៉ង់: `{order_id}`\n"
            f"⏰ ម៉ោង: {time_now}\n"
            f"--------------------------\n"
            f"📦 **មុខម្ហូប:**\n{items_detail_msg}\n"
            f"💰 **សរុប: ${total:.2f}**\n"
            f"--------------------------\n"
            f"📍 ចម្ងាយ: {distance_text}\n"
            f"🔗 ទីតាំងភ្ញៀវ: [មើលលើផែនទី]({map_link})"
        )

        # 5. ផ្ញើទៅ Telegram (ប្រើ Timeout ដើម្បីការពារការគាំង)
        try:
            tel_response = requests.post(
                f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage",
                json={"chat_id": CHAT_ID, "text": telegram_msg, "parse_mode": "Markdown"},
                timeout=10
            )
            tel_response.raise_for_status()
        except Exception as tel_err:
            print(f"❌ Telegram Error: {tel_err}")
            return jsonify({"success": False, "error": "មិនអាចបញ្ជូនដំណឹងទៅអ្នកលក់បានទេ"}), 500

        # 6. កត់ត្រាចូល Excel
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            ws.append([
                order_id, 
                queue_number, 
                time_now, 
                items_for_excel.rstrip(', '), 
                total, 
                round(distance_m, 2), 
                map_link
            ])
            wb.save(EXCEL_FILE)
        except Exception as excel_err:
            print(f"❌ Excel Save Error: {excel_err}")

        return jsonify({
            "success": True, 
            "order_id": order_id,
            "message": f"ការកុម្ម៉ង់បានជោគជ័យ! លេខរង់ចាំរបស់អ្នកគឺ: {queue_number}"
        })

    except Exception as e:
        print(f"❌ Global Error: {e}")
        return jsonify({"success": False, "error": "មានបញ្ហាបច្ចេកទេសនៅលើ Server"}), 500


# ================== ADMIN PANEL ROUTES ==================

@app.route("/admin")
@login_required
def admin_panel():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # អាន header
        headers = [cell.value for cell in ws[1]]
        
        # អាន data rows
        orders = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # ប្រសិនបើមាន Order ID
                orders.append(dict(zip(headers, row)))
        
        # បញ្ច្រាស់លំដាប់ (ថ្មីជាងមុន)
        orders.reverse()
        
        return render_template_string(ADMIN_TEMPLATE, orders=orders, total_orders=len(orders))
    
    except Exception as e:
        return f"Error loading orders: {e}", 500


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('admin_panel'))
        else:
            return render_template_string(LOGIN_TEMPLATE, error="Username ឬ Password មិនត្រឹមត្រូវ!")
    
    return render_template_string(LOGIN_TEMPLATE)


@app.route("/admin/logout")
def admin_logout():
    session.pop('logged_in', None)
    return redirect(url_for('admin_login'))


# ================== HTML TEMPLATES ==================

LOGIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="km">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Admin Login - Tube Coffee</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Khmer OS Battambang', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }
        .login-container {
            background: white;
            padding: 40px;
            border-radius: 15px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            width: 100%;
            max-width: 400px;
        }
        .logo {
            text-align: center;
            font-size: 48px;
            margin-bottom: 10px;
        }
        h2 {
            text-align: center;
            color: #333;
            margin-bottom: 30px;
        }
        .form-group {
            margin-bottom: 20px;
        }
        label {
            display: block;
            margin-bottom: 8px;
            color: #555;
            font-weight: bold;
        }
        input {
            width: 100%;
            padding: 12px;
            border: 2px solid #ddd;
            border-radius: 8px;
            font-size: 14px;
            transition: border 0.3s;
        }
        input:focus {
            outline: none;
            border-color: #667eea;
        }
        button {
            width: 100%;
            padding: 14px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 8px;
            font-size: 16px;
            font-weight: bold;
            cursor: pointer;
            transition: transform 0.2s;
        }
        button:hover {
            transform: translateY(-2px);
        }
        .error {
            background: #fee;
            color: #c33;
            padding: 12px;
            border-radius: 8px;
            margin-bottom: 20px;
            text-align: center;
        }
    </style>
</head>
<body>
    <div class="login-container">
        <div class="logo">☕</div>
        <h2>Tube Coffee Admin</h2>
        {% if error %}
        <div class="error">{{ error }}</div>
        {% endif %}
        <form method="POST">
            <div class="form-group">
                <label>Username:</label>
                <input type="text" name="username" required autofocus>
            </div>
            <div class="form-group">
                <label>Password:</label>
                <input type="password" name="password" required>
            </div>
            <button type="submit">🔐 ចូលប្រើប្រាស់</button>
        </form>
    </div>
</body>
</html>
"""

ADMIN_TEMPLATE = """
<!DOCTYPE html>
<html lang="km">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Admin Panel - Tube Coffee</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: 'Khmer OS Battambang', Arial, sans-serif;
            background: #f5f5f5;
        }
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            display: flex;
            justify-content: space-between;
            align-items: center;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        .header h1 {
            font-size: 24px;
        }
        .logout-btn {
            background: rgba(255,255,255,0.2);
            color: white;
            padding: 10px 20px;
            border-radius: 8px;
            text-decoration: none;
            transition: background 0.3s;
        }
        .logout-btn:hover {
            background: rgba(255,255,255,0.3);
        }
        .stats {
            display: flex;
            gap: 20px;
            padding: 20px;
            max-width: 1200px;
            margin: 0 auto;
        }
        .stat-card {
            flex: 1;
            background: white;
            padding: 20px;
            border-radius: 12px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        .stat-card h3 {
            color: #666;
            font-size: 14px;
            margin-bottom: 10px;
        }
        .stat-card .number {
            font-size: 32px;
            font-weight: bold;
            color: #667eea;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
        }
        table {
            width: 100%;
            background: white;
            border-radius: 12px;
            overflow: hidden;
            box-shadow: 0 2px 10px rgba(0,0,0,0.1);
        }
        th {
            background: #667eea;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: bold;
        }
        td {
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }
        tr:hover {
            background: #f9f9f9;
        }
        .map-link {
            color: #667eea;
            text-decoration: none;
        }
        .map-link:hover {
            text-decoration: underline;
        }
        .no-orders {
            text-align: center;
            padding: 60px 20px;
            color: #999;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>☕ Tube Coffee - Admin Panel</h1>
        <a href="/admin/logout" class="logout-btn">🚪 ចាកចេញ</a>
    </div>
    
    <div class="stats">
        <div class="stat-card">
            <h3>សរុបការកុម្ម៉ង់ទាំងអស់</h3>
            <div class="number">{{ total_orders }}</div>
        </div>
        <div class="stat-card">
            <h3>ការកុម្ម៉ង់ថ្ងៃនេះ</h3>
            <div class="number">-</div>
        </div>
        <div class="stat-card">
            <h3>ចំណូលសរុប</h3>
            <div class="number">-</div>
        </div>
    </div>
    
    <div class="container">
        {% if orders %}
        <table>
            <thead>
                <tr>
                    <th>Order ID</th>
                    <th>លេខរង់ចាំ</th>
                    <th>ម៉ោង</th>
                    <th>មុខម្ហូប</th>
                    <th>សរុប ($)</th>
                    <th>ចម្ងាយ (m)</th>
                    <th>ផែនទី</th>
                </tr>
            </thead>
            <tbody>
                {% for order in orders %}
                <tr>
                    <td><strong>{{ order['Order ID'] }}</strong></td>
                    <td>{{ order['Queue Number'] }}</td>
                    <td>{{ order['Time'] }}</td>
                    <td>{{ order['Items'] }}</td>
                    <td><strong>${{ order['Total ($)'] }}</strong></td>
                    <td>{{ order['Distance (m)'] }}</td>
                    <td>
                        {% if order['Map Link'] != 'N/A' %}
                        <a href="{{ order['Map Link'] }}" target="_blank" class="map-link">📍 មើល</a>
                        {% else %}
                        -
                        {% endif %}
                    </td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
        {% else %}
        <div class="no-orders">
            <h2>📋 មិនទាន់មានការកុម្ម៉ង់ទេ</h2>
        </div>
        {% endif %}
    </div>
</body>
</html>
"""

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)